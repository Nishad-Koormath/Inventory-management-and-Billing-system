from django.shortcuts import render, redirect
from .forms import StockTransactionForm
from django.core.exceptions import ValidationError
from product_app.models import Product, Category
from .models import StockTransaction
from django.utils.dateparse import parse_date
from django.db.models import Sum
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.utils.dateparse import parse_date





# Create your views here.

# stock
def stock_transaction(request):
    if request.method == 'POST':
        form = StockTransactionForm(request.POST)
        if form.is_valid():
            try:
                form.save()
                return redirect('product_list')
            except ValidationError as e:
                form.add_error(None, e)
    else:
        form = StockTransactionForm()
    return render(request, 'inventory_app/stock_transaction.html', {'form': form})

def stock_transaction_list(request):
    products = Product.objects.all()
    categories = Category.objects.all()
    transactions = StockTransaction.objects.select_related('product').order_by('-timestamp')
    
    product_id = request.GET.get('product')
    category_id = request.GET.get('category')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if category_id:
        transactions = transactions.filter(product__category__id=category_id)
    if product_id:
        transactions = transactions.filter(product__id = product_id )
    if start_date:
        transactions = transactions.filter(timestamp__date__gte=parse_date(start_date))
    if end_date:
        transactions = transactions.filter(timestamp__date__lte=parse_date(end_date))
        
    total_in = transactions.filter(transaction_types = 'IN').aggregate(Sum('quantity'))['quantity__sum'] or 0
    total_out = transactions.filter(transaction_types = 'OUT').aggregate(Sum('quantity'))['quantity__sum'] or 0
    
        
    return render(request, 'inventory_app/stock_transaction_list.html', {
        'transactions': transactions,
        'products': products,
        'categories': categories,
        'total_in': total_in,
        'total_out': total_out,
        })
    
    
def stock_report(request):
    categories = Category.objects.all()
    products = Product.objects.all()
    
    selected_category = request.GET.get('category')
    selected_product = request.GET.get('product')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    export = request.GET.get('export')
    
    if selected_category == 'None':
        selected_category = None
    if selected_product == 'None':
        selected_product = None
    if start_date == 'None':
        start_date = None
    if end_date == 'None':
        end_date = None
    
    filtered_product = products
    
    if selected_category:
        filtered_product = filtered_product.filter(category_id = selected_category)
    if selected_product:
        filtered_product = filtered_product.filter(id = selected_product)
    
    report_data = []
    
    for product in filtered_product:
        transactions = StockTransaction.objects.filter(product=product).order_by('timestamp')
        
        if start_date:
            transactions = transactions.filter(timestamp__date__gte = start_date)
        if end_date:
            transactions = transactions.filter(timestamp__date_lte = end_date)
        
        balance = 0
        product_report = []
        
        for tx in transactions:
            qty_in = tx.quantity if tx.transaction_types == 'IN' else 0
            qty_out = tx.quantity if tx.transaction_types == 'OUT' else 0
            balance += qty_in - qty_out
            
            product_report.append({
                'date': tx.timestamp,
                'product': product.name,
                'in': qty_in,
                'out': qty_out, 
                'balance': balance,
            })
            
        if product_report:
            report_data.extend(product_report)
        
    if export == 'excel':
        wb = Workbook()
        ws = wb.active
        ws.title = "Stock report"
        
        headers = ['Date', 'Product', 'In', 'Out', "Balance"]
        ws.append(headers)
        
        for row in report_data:
            ws.append([
                row['date'].strftime('%d-%m-%Y %H:%M'),
                row['product'],
                row['in'],
                row['out'],
                row['balance']
            ])
        
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2
            
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=stock_report.xlsx'
        wb.save(response)
        return response
        
    context = {
        'categories': categories,
        'products': products,
        'report_data': report_data,
        'selected_category': selected_category,
        'selected_product': selected_product,
        'start_date': start_date,
        'end_date': end_date,
    }
    return render(request, 'inventory_app/stock_report.html', context)